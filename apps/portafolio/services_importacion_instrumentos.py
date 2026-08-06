"""Importador estructurado y reutilizable de instrumentos Excel de Portafolio."""

import hashlib
import json
from pathlib import Path

from django.core.files import File
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import (
    CalculadoraInstrumento,
    CategoriaDocumento,
    Documento,
    ImportacionInstrumento,
    Instrumento,
    PreguntaInstrumento,
    RevisionInstrumento,
)


HOJAS_REQUERIDAS = {
    'INSTRUMENTO',
    'PREGUNTAS',
    'CALCULADORA_SISTEMA',
    'CASOS_PRUEBA',
}

TIPOS = {
    'escala',
    'si_no',
    'opcion_unica',
    'opcion_multiple',
    'texto_libre',
}


def _filas(hoja):
    return list(
        hoja.iter_rows(
            values_only=True,
        )
    )


def _tabla(filas, encabezado):
    for indice, fila in enumerate(filas):
        if any(
            str(valor or '').strip() == encabezado
            for valor in fila
        ):
            columnas = [
                str(valor or '').strip()
                for valor in fila
            ]

            return [
                {
                    columnas[i]: valor
                    for i, valor in enumerate(actual)
                    if i < len(columnas) and columnas[i]
                }
                for actual in filas[indice + 1:]
                if any(
                    valor is not None and str(valor).strip()
                    for valor in actual
                )
            ]

    return []


def leer_excel(ruta):
    ruta = Path(ruta)

    try:
        contenido = ruta.read_bytes()
        libro = load_workbook(
            ruta,
            read_only=True,
            data_only=False,
        )
    except Exception as error:
        raise ValidationError(
            f'{ruta.name}: no fue posible leer el archivo.'
        ) from error

    faltantes = HOJAS_REQUERIDAS - set(libro.sheetnames)

    if faltantes:
        raise ValidationError(
            f'{ruta.name}: faltan hojas requeridas: '
            f'{", ".join(sorted(faltantes))}.'
        )

    instrumento_filas = _filas(
        libro['INSTRUMENTO']
    )

    valores = {
        str(f[0]).strip(): f[1]
        for f in instrumento_filas
        if (
            len(f) > 1
            and f[0]
            and str(f[0]).strip() not in {'Campo'}
        )
    }

    preguntas = _tabla(
        _filas(libro['PREGUNTAS']),
        'instrumento_clave',
    )

    calculadora_filas = _filas(
        libro['CALCULADORA_SISTEMA']
    )

    calculadora = {
        str(f[0]).strip(): f[1]
        for f in calculadora_filas
        if (
            len(f) > 1
            and f[0]
            and str(f[0]).strip() not in {'Campo'}
        )
    }

    casos = _tabla(
        _filas(libro['CASOS_PRUEBA']),
        'Caso',
    )

    if not preguntas:
        raise ValidationError(
            f'{ruta.name}: PREGUNTAS no contiene filas.'
        )

    if (
        not calculadora.get('clave_calculadora')
        or not calculadora.get('version_regla')
    ):
        raise ValidationError(
            f'{ruta.name}: CALCULADORA_SISTEMA requiere '
            'clave_calculadora y version_regla.'
        )

    return {
        'ruta': ruta,
        'contenido': contenido,
        'huella': hashlib.sha256(contenido).hexdigest(),
        'instrumento': valores,
        'preguntas': preguntas,
        'calculadora': calculadora,
        'calculadora_filas': calculadora_filas,
        'casos': casos,
    }


def validar(datos):
    preguntas = datos['preguntas']
    claves = set()
    ordenes = set()

    for numero, pregunta in enumerate(
        preguntas,
        2,
    ):
        clave, orden, texto, tipo = (
            pregunta.get('pregunta_clave'),
            pregunta.get('orden'),
            pregunta.get('texto'),
            str(
                pregunta.get('tipo_respuesta') or ''
            ).strip(),
        )

        if not clave or clave in claves:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'clave duplicada o vacía.'
            )

        if not isinstance(orden, int) or orden in ordenes:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'orden duplicado o inválido.'
            )

        if not texto:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'texto vacío.'
            )

        if tipo not in TIPOS:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                f'tipo no soportado: {tipo}.'
            )

        try:
            opciones = json.loads(
                pregunta.get('opciones_json') or 'null'
            )
        except json.JSONDecodeError as error:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'opciones_json inválido.'
            ) from error

        if opciones is not None and (
            not isinstance(opciones, list)
            or any(
                not isinstance(o, dict)
                or 'valor' not in o
                or 'etiqueta' not in o
                for o in opciones
            )
        ):
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'opciones_json inválido.'
            )

        claves.add(clave)
        ordenes.add(orden)

    primera = preguntas[0]

    for campo in (
        'instrumento_clave',
        'instrumento_nombre',
        'version',
    ):
        if not primera.get(campo):
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS requiere {campo}.'
            )

    return datos


def _estado(valor):
    texto = str(valor or '').upper()

    if 'BLOQUE' in texto:
        return CalculadoraInstrumento.Estado.BLOQUEADA

    if 'NO ACTIVA' in texto or 'DIAGN' in texto:
        return CalculadoraInstrumento.Estado.NO_DIAGNOSTICA

    if 'ORIENTAT' in texto:
        return CalculadoraInstrumento.Estado.ORIENTATIVA

    return CalculadoraInstrumento.Estado.ACTIVA


def _estado_calculadora_por_instrumento(clave, calculadora):
    """Conserva la interpretación orientativa de las variantes autorizadas."""
    estado = _estado(
        calculadora.get('estado_calculadora')
        or calculadora.get('estado_puntaje')
        or calculadora.get('estado_interpretacion')
    )
    if clave in {
        'dass-21-adolescentes',
        'rse-autoestima',
    } and _estado(calculadora.get('estado_interpretacion')) == (
        CalculadoraInstrumento.Estado.ORIENTATIVA
    ):
        return CalculadoraInstrumento.Estado.ORIENTATIVA
    return estado


def _definicion(datos):
    clave = datos['preguntas'][0]['instrumento_clave']
    clase = 'conteo'

    if clave.startswith('dass-21'):
        clase = 'subescalas_multiplicadas'
    elif clave == 'rse-autoestima':
        clase = 'suma_recodificada'
    elif 'plutchik' in clave:
        clase = 'suma_criticos'

    return {
        'algoritmo': clase,
        'campos': datos['calculadora'],
        'tablas': datos['calculadora_filas'],
        'casos_prueba': datos['casos'],
        'requiere_respuestas_completas': str(
            datos['calculadora'].get(
                'requiere_respuestas_completas',
                '',
            )
        ).lower() in {
            'true',
            'sí',
            'si',
        },
        'sin_acciones_automaticas': True,
    }


def ejecutar_calculadora(
    definicion,
    respuestas,
    edad=None,
):
    if (
        definicion.get('estado')
        == CalculadoraInstrumento.Estado.BLOQUEADA
    ):
        return {
            'bloqueada': True,
        }

    algoritmo = definicion['algoritmo']

    valores = {
        int(k): int(v)
        for k, v in respuestas.items()
        if v not in (
            None,
            '',
        )
    }

    if algoritmo == 'subescalas_multiplicadas':
        grupos = {
            'Depresión': [
                3,
                5,
                10,
                13,
                16,
                17,
                21,
            ],
            'Ansiedad': [
                2,
                4,
                7,
                9,
                15,
                19,
                20,
            ],
            'Estrés': [
                1,
                6,
                8,
                11,
                12,
                14,
                18,
            ],
        }

        detalle = {
            n: sum(
                valores.get(i, 0)
                for i in items
            ) * 2
            for n, items in grupos.items()
        }

        return {
            'puntaje_total': sum(detalle.values()),
            'detalle': detalle,
        }

    if algoritmo == 'suma_recodificada':
        inversos = {
            2,
            5,
            8,
            9,
            10,
        }

        total = sum(
            5 - v if i in inversos else v
            for i, v in valores.items()
        )

        nivel = (
            'Autoestima elevada'
            if total >= 30
            else (
                'Autoestima media'
                if total >= 26
                else (
                    'SIN RANGO DEFINIDO'
                    if total == 25
                    else 'Autoestima baja'
                )
            )
        )

        return {
            'puntaje_total': total,
            'interpretacion': nivel,
            'detalle': {},
        }

    if algoritmo == 'suma_criticos':
        total = sum(valores.values())

        return {
            'puntaje_total': total,
            'interpretacion': (
                'Presencia de riesgo; evaluación inmediata'
                if total >= 6
                else 'No se detecta riesgo mediante este tamizaje'
            ),
            'focos': [
                i
                for i in (
                    13,
                    14,
                    15,
                )
                if valores.get(i) == 1
            ],
        }

    return {
        'puntaje_total': sum(valores.values()),
        'detalle': {
            'conteos': sum(valores.values()),
            'edad': edad,
        },
    }


def importar(ruta, dry_run=False):
    datos = validar(
        leer_excel(ruta)
    )

    primera = datos['preguntas'][0]
    clave = primera['instrumento_clave']
    version = str(primera['version'])
    definicion = _definicion(datos)

    estado = _estado_calculadora_por_instrumento(
        clave,
        datos['calculadora'],
    )

    reporte = {
        'archivo': datos['ruta'].name,
        'clave': clave,
        'version': version,
        'preguntas': len(datos['preguntas']),
        'casos': len(datos['casos']),
        'estado_calculadora': estado,
        'decision': 'importable',
        'dry_run': dry_run,
    }

    existente = Instrumento.objects.filter(
        clave=clave
    ).first()

    if (
        existente
        and hasattr(existente, 'importacion')
        and existente.importacion.huella_contenido == datos['huella']
    ):
        reporte['decision'] = 'sin cambios'
        return reporte

    if existente and (
        existente.aplicaciones_intera.exists()
        or existente.preguntas.filter(
            respuestas_intera__isnull=False
        ).exists()
    ):
        raise ValidationError(
            f'{clave}: instrumento utilizado; no se sobrescribe.'
        )

    if dry_run:
        return reporte

    with transaction.atomic():
        categoria, _ = CategoriaDocumento.objects.get_or_create(
            nombre='Instrumento'
        )

        documento = (
            ImportacionInstrumento.objects
            .filter(
                huella_contenido=datos['huella']
            )
            .select_related('documento')
            .values_list(
                'documento',
                flat=True,
            )
            .first()
        )

        documento = (
            Documento.objects.filter(
                pk=documento
            ).first()
            if documento
            else None
        )

        if not documento:
            with datos['ruta'].open('rb') as archivo:
                documento = Documento(
                    nombre=datos['ruta'].stem,
                    categoria=categoria,
                    version=version,
                    descripcion=(
                        'Documento origen importado por Portafolio.'
                    ),
                )

                documento.archivo.save(
                    datos['ruta'].name,
                    File(archivo),
                    save=False,
                )

                documento.save()

        instrumento = existente or Instrumento(
            clave=clave
        )

        instrumento.nombre = primera['instrumento_nombre']
        instrumento.version = version
        instrumento.documento_origen = documento
        instrumento.descripcion = str(
            datos['instrumento'].get('descripcion') or ''
        )
        instrumento.instrucciones = str(
            datos['instrumento'].get('instrucciones') or ''
        )
        instrumento.full_clean()
        instrumento.save()

        ImportacionInstrumento.objects.update_or_create(
            instrumento=instrumento,
            defaults={
                'documento': documento,
                'huella_contenido': datos['huella'],
                'metadatos': {
                    **datos['instrumento'],
                    'variante': (
                        primera.get('variante')
                        or datos['instrumento'].get('variante')
                    ),
                    'poblacion': primera.get('poblacion'),
                    'edad_min': primera.get('edad_min'),
                    'edad_max': primera.get('edad_max'),
                },
            },
        )

        instrumento.preguntas.all().delete()

        PreguntaInstrumento.objects.bulk_create(
            [
                PreguntaInstrumento(
                    instrumento=instrumento,
                    orden=p['orden'],
                    clave=p['pregunta_clave'],
                    texto=p['texto'],
                    tipo_respuesta=p['tipo_respuesta'],
                    opciones=json.loads(
                        p.get('opciones_json') or 'null'
                    ),
                    requerida=bool(p.get('requerida')),
                    condicion_visibilidad=(
                        p.get('visibilidad') or None
                    ),
                )
                for p in datos['preguntas']
            ]
        )

        RevisionInstrumento.objects.update_or_create(
            instrumento=instrumento,
            version=version,
            defaults={
                'estructura': {
                    'metadatos': instrumento.importacion.metadatos,
                    'preguntas': list(
                        instrumento.preguntas.values(
                            'orden',
                            'clave',
                            'texto',
                            'tipo_respuesta',
                            'opciones',
                            'requerida',
                        )
                    ),
                },
            },
        )

        CalculadoraInstrumento.objects.update_or_create(
            instrumento=instrumento,
            clave=datos['calculadora']['clave_calculadora'],
            version_regla=str(
                datos['calculadora']['version_regla']
            ),
            defaults={
                'estado': estado,
                'definicion': {
                    **definicion,
                    'estado': estado,
                },
                'huella_contenido': datos['huella'],
            },
        )

    return reporte
