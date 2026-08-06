import re

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import PreguntaInstrumento


_PREGUNTA = re.compile(
    r'^\s*(\d+)\s*[\.)]\s*(.+)$'
)


def _tipo_y_opciones(identificador):
    identificador = identificador.lower()

    if 'radio' in identificador:
        return PreguntaInstrumento.Tipo.SI_NO, [
            {
                'valor': '1',
                'etiqueta': 'Sí',
            },
            {
                'valor': '0',
                'etiqueta': 'No',
            },
        ]

    if 'checkbox' in identificador:
        return (
            PreguntaInstrumento.Tipo.OPCION_MULTIPLE,
            [],
        )

    if (
        'textarea' in identificador
        or 'text' in identificador
    ):
        return (
            PreguntaInstrumento.Tipo.TEXTO_LIBRE,
            None,
        )

    return (
        PreguntaInstrumento.Tipo.TEXTO_LIBRE,
        None,
    )


def importar_preguntas_desde_documento(instrumento):
    """Importa preguntas desde el Excel almacenado en el Documento origen.

    El formato reconoce encabezados de formulario como ``1. Pregunta|radio-57``.
    El archivo nunca se copia: se abre mediante el almacenamiento de Portafolio.
    """

    documento = instrumento.documento_origen

    if not documento or not documento.archivo:
        raise ValidationError(
            'El instrumento necesita un Documento origen en Portafolio.'
        )

    if documento.tipo_archivo.lower() != 'xlsx':
        raise ValidationError(
            'Solo se pueden importar documentos Excel .xlsx.'
        )

    try:
        with documento.archivo.open('rb') as archivo:
            libro = load_workbook(
                archivo,
                read_only=True,
                data_only=False,
            )

            hoja = max(
                libro.worksheets,
                key=lambda actual: sum(
                    '|' in str(valor)
                    for valor in next(
                        actual.iter_rows(
                            min_row=1,
                            max_row=1,
                            values_only=True,
                        ),
                        (),
                    )
                ),
            )

            encabezados = next(
                hoja.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                (),
            )

    except Exception as error:
        raise ValidationError(
            'No fue posible leer el Excel del Documento origen.'
        ) from error

    preguntas = []
    ordenes = set()

    for encabezado in encabezados:
        texto_completo = str(
            encabezado or ''
        ).strip()

        if '|' not in texto_completo:
            continue

        texto, identificador = (
            parte.strip()
            for parte in texto_completo.split(
                '|',
                1,
            )
        )

        coincidencia = _PREGUNTA.match(texto)

        if not coincidencia:
            continue

        orden = int(
            coincidencia.group(1)
        )
        texto = coincidencia.group(2).strip()

        if orden in ordenes:
            continue

        ordenes.add(orden)

        tipo, opciones = _tipo_y_opciones(
            identificador
        )

        preguntas.append(
            PreguntaInstrumento(
                instrumento=instrumento,
                orden=orden,
                texto=texto,
                clave=identificador,
                tipo_respuesta=tipo,
                opciones=opciones,
                requerida=True,
            )
        )

    if not preguntas:
        raise ValidationError(
            (
                'El Excel no contiene encabezados de preguntas '
                'compatibles para importar.'
            )
        )

    if instrumento.preguntas.filter(
        respuestas_intera__isnull=False
    ).exists():
        raise ValidationError(
            (
                'No se pueden reemplazar preguntas porque el '
                'instrumento ya tiene respuestas registradas.'
            )
        )

    with transaction.atomic():
        instrumento.preguntas.all().delete()

        PreguntaInstrumento.objects.bulk_create(
            preguntas
        )

    return len(preguntas)