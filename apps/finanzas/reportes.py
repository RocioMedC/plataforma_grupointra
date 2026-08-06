"""Los tres documentos de la pantalla de Reportes: Estado de resultados
(PDF), Flujo de efectivo (XLSX) y Concentrado de donativos (PDF).

Todos reciben el mismo rango de fechas — el que se elige en el modal de la
pantalla — donde `desde=None, hasta=None` significa **todo el historial**.
Los cálculos usan `totales.py`, el mismo criterio de dinero real que el
tablero, para que el documento descargado y la pantalla nunca discrepen.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import Max, Min, Sum
from django.http import HttpResponse

from .models import CategoriaEgreso, Donativo, Egreso, Ingreso
from .totales import donativos_efectivos, egresos_efectivos, ingresos_efectivos

MESES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]
MESES_ABREV = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']


# ===== Rango de fechas =====

def rango_pedido(request):
    """Lee el rango que eligió el usuario en el modal de Reportes. El modal
    manda `periodo=todo` o `periodo=rango`; solo en el segundo caso se miran
    las fechas, así que "Todo el historial" no depende de que los campos de
    fecha hayan quedado vacíos."""
    if request.GET.get('periodo') != 'rango':
        return None, None
    desde = _fecha(request.GET.get('desde'))
    hasta = _fecha(request.GET.get('hasta'))
    # Si vinieron al revés se enderezan, en vez de entregar un documento
    # vacío sin explicar por qué.
    if desde and hasta and desde > hasta:
        desde, hasta = hasta, desde
    return desde, hasta


def _fecha(valor):
    if not valor:
        return None
    try:
        return date.fromisoformat(valor)
    except ValueError:
        return None


def en_rango(queryset, desde, hasta):
    if desde:
        queryset = queryset.filter(fecha__gte=desde)
    if hasta:
        queryset = queryset.filter(fecha__lte=hasta)
    return queryset


def etiqueta_periodo(desde, hasta):
    """Texto que va en el encabezado del documento y explica qué se incluyó."""
    if desde and hasta:
        return f'Del {_dia(desde)} al {_dia(hasta)}'
    if desde:
        return f'Del {_dia(desde)} en adelante'
    if hasta:
        return f'Hasta el {_dia(hasta)}'
    return 'Todo el historial'


def _dia(valor):
    return f'{valor.day} de {MESES[valor.month].lower()} de {valor.year}'


def sufijo_archivo(desde, hasta):
    """Parte variable del nombre del archivo, para que dos descargas de
    rangos distintos no se pisen en la carpeta de Descargas."""
    if desde and hasta:
        return f'{desde.isoformat()}_a_{hasta.isoformat()}'
    if desde:
        return f'desde_{desde.isoformat()}'
    if hasta:
        return f'hasta_{hasta.isoformat()}'
    return 'historico'


# ===== Estado de resultados =====

def estado_de_resultados(desde=None, hasta=None):
    """Renglones del Estado de resultados de un rango. Lo usan la pantalla
    (`reportes_view`) y el PDF, para que sean literalmente el mismo cálculo."""
    ingresos = en_rango(Ingreso.objects.all(), desde, hasta)
    egresos = en_rango(Egreso.objects.all(), desde, hasta)
    donativos = en_rango(Donativo.objects.all(), desde, hasta)

    servicios = ingresos_efectivos(ingresos)
    donado = donativos_efectivos(donativos)

    renta = (
        egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.RENTA))
        + egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.SERVICIOS))
    )
    nomina_admin = egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.NOMINA_ADMIN))
    nomina_terapeutas = egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.NOMINA_TERAPEUTAS))
    nomina_academia = egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.NOMINA_ACADEMIA))
    insumos = egresos_efectivos(egresos.filter(categoria=Egreso.Categoria.INSUMOS))

    # "Otros egresos": cualquier categoría agregada desde Configuración que
    # no sea una de las fijas de arriba. Así el total siempre cuadra con la
    # suma real, sin importar cuántas categorías nuevas agregue el usuario.
    categorias_fijas = [c[0] for c in Egreso.Categoria.choices]
    otros_qs = egresos.exclude(categoria__in=categorias_fijas)
    otros = egresos_efectivos(otros_qs)

    total_ingresos = servicios + donado
    total_egresos = renta + nomina_admin + nomina_terapeutas + nomina_academia + insumos + otros

    return {
        'ingresos_servicios': servicios,
        'donativos': donado,
        'total_ingresos': total_ingresos,
        'nomina_admin': nomina_admin,
        'nomina_terapeutas': nomina_terapeutas,
        'nomina_academia': nomina_academia,
        'renta': renta,
        'insumos': insumos,
        'otros': otros,
        'otros_detalle': _detalle_otros(otros_qs),
        'total_egresos': total_egresos,
        'resultado': total_ingresos - total_egresos,
        'ingresos_por_concepto': _detalle_ingresos(ingresos),
    }


def _nombres_de_categoria():
    """Etiqueta legible de cada categoría de egreso, juntando las fijas del
    modelo con las que el usuario agregó desde Configuración."""
    nombres = dict(Egreso.Categoria.choices)
    for cat in CategoriaEgreso.objects.all():
        nombres.setdefault(cat.nombre, cat.nombre)
    return nombres


def _detalle_otros(queryset):
    """Desglose del renglón "Otros egresos" por categoría. La pantalla solo
    muestra el total; el PDF sí lista de qué se compone."""
    nombres = _nombres_de_categoria()
    filas = []
    for fila in (
        queryset.filter(estatus=Egreso.Estatus.PAGADO)
        .values('categoria').annotate(total=Sum('monto')).order_by('-total')
    ):
        filas.append({
            'nombre': nombres.get(fila['categoria'], fila['categoria']),
            'total': fila['total'] or Decimal('0'),
        })
    return filas


def _detalle_ingresos(queryset):
    """Ingresos efectivos agrupados por concepto, para el anexo del PDF.
    Mismo criterio que `totales.ingresos_efectivos`: de un Parcial se toma lo
    cobrado, y un Pendiente no aparece."""
    nombres = dict(Ingreso.Concepto.choices)
    totales = {}
    for fila in queryset.filter(estatus=Ingreso.Estatus.PAGADO).values('concepto').annotate(total=Sum('monto')):
        totales[fila['concepto']] = totales.get(fila['concepto'], Decimal('0')) + (fila['total'] or Decimal('0'))
    for fila in queryset.filter(estatus=Ingreso.Estatus.PARCIAL).values('concepto').annotate(total=Sum('monto_pagado')):
        totales[fila['concepto']] = totales.get(fila['concepto'], Decimal('0')) + (fila['total'] or Decimal('0'))
    filas = [
        {'nombre': nombres.get(clave, clave), 'total': monto}
        for clave, monto in totales.items() if monto
    ]
    filas.sort(key=lambda f: f['total'], reverse=True)
    return filas


# ===== Flujo de efectivo =====

def _limites_de_datos(desde, hasta):
    """Primer y último mes con movimientos dentro del rango. Sirve para que
    "Todo el historial" no tenga que recorrer meses vacíos."""
    if desde and hasta:
        return desde, hasta
    minimos, maximos = [], []
    for modelo in (Ingreso, Egreso, Donativo):
        limites = en_rango(modelo.objects.all(), desde, hasta).aggregate(minima=Min('fecha'), maxima=Max('fecha'))
        if limites['minima']:
            minimos.append(limites['minima'])
        if limites['maxima']:
            maximos.append(limites['maxima'])
    if not minimos:
        return None, None
    return desde or min(minimos), hasta or max(maximos)


def _meses_entre(inicio, fin):
    meses = []
    anio, mes = inicio.year, inicio.month
    while (anio, mes) <= (fin.year, fin.month):
        meses.append((anio, mes))
        anio, mes = (anio + 1, 1) if mes == 12 else (anio, mes + 1)
    return meses


def flujo_mensual(desde=None, hasta=None):
    """Entradas y salidas mes por mes, con el saldo acumulado. Solo dinero
    que se movió de verdad (mismo criterio que el resto del módulo)."""
    inicio, fin = _limites_de_datos(desde, hasta)
    if not inicio:
        return []

    ingresos = en_rango(Ingreso.objects.all(), desde, hasta)
    egresos = en_rango(Egreso.objects.all(), desde, hasta)
    donativos = en_rango(Donativo.objects.all(), desde, hasta)

    filas = []
    acumulado = Decimal('0')
    for anio, mes in _meses_entre(inicio, fin):
        cobrado = ingresos_efectivos(ingresos.filter(fecha__year=anio, fecha__month=mes))
        donado = donativos_efectivos(donativos.filter(fecha__year=anio, fecha__month=mes))
        pagado = egresos_efectivos(egresos.filter(fecha__year=anio, fecha__month=mes))
        entradas = cobrado + donado
        acumulado += entradas - pagado
        filas.append({
            'etiqueta': f'{MESES[mes]} {anio}',
            'anio': anio,
            'mes': mes,
            'servicios': cobrado,
            'donativos': donado,
            'entradas': entradas,
            'salidas': pagado,
            'neto': entradas - pagado,
            'acumulado': acumulado,
        })
    return filas


def movimientos_de_efectivo(desde=None, hasta=None):
    """Detalle que sustenta el resumen mensual: cada movimiento que sí es
    dinero (de un ingreso Parcial se toma lo cobrado, no el monto total, y
    los Pendientes no aparecen)."""
    filas = []
    for i in en_rango(Ingreso.objects.select_related('terapeuta'), desde, hasta):
        if i.estatus == Ingreso.Estatus.PAGADO:
            monto = i.monto
        elif i.estatus == Ingreso.Estatus.PARCIAL:
            monto = i.monto_pagado
        else:
            continue
        if not monto:
            continue
        filas.append({
            'fecha': i.fecha, 'flujo': 'Entrada', 'tipo': 'Ingreso',
            'unidad': i.get_unidad_display(), 'concepto': i.get_concepto_display(),
            'persona': i.persona or (str(i.terapeuta) if i.terapeuta else ''),
            'estatus': i.get_estatus_display(), 'monto': monto,
        })
    for d in en_rango(Donativo.objects.all(), desde, hasta).exclude(estatus_cfdi=Donativo.EstatusCFDI.CANCELADO):
        filas.append({
            'fecha': d.fecha, 'flujo': 'Entrada', 'tipo': 'Donativo',
            # Los donativos son institucionales, no de una unidad (ver tablero_view).
            'unidad': '', 'concepto': f'Donativo {d.get_tipo_display().lower()}',
            'persona': d.donante_nombre, 'estatus': d.get_estatus_cfdi_display(), 'monto': d.monto,
        })
    for e in en_rango(Egreso.objects.filter(estatus=Egreso.Estatus.PAGADO), desde, hasta):
        filas.append({
            'fecha': e.fecha, 'flujo': 'Salida', 'tipo': 'Egreso',
            'unidad': e.get_unidad_display(), 'concepto': e.concepto,
            'persona': e.persona, 'estatus': e.get_estatus_display(), 'monto': -e.monto,
        })
    filas.sort(key=lambda f: f['fecha'])
    return filas


def flujo_efectivo_xlsx(desde=None, hasta=None, generado_por=''):
    """Arma el libro de Excel del Flujo de efectivo: una hoja con el resumen
    mensual y otra con el detalle que lo sustenta."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    azul = '1B2C4F'
    encabezado = Font(bold=True, color='FFFFFF', size=10)
    relleno = PatternFill('solid', fgColor=azul)
    borde_abajo = Border(bottom=Side(style='thin', color='E6EAF2'))
    moneda = '"$"#,##0.00;[Red]-"$"#,##0.00'

    libro = Workbook()

    resumen = libro.active
    resumen.title = 'Resumen mensual'
    resumen['A1'] = 'Flujo de efectivo · Grupo Intra'
    resumen['A1'].font = Font(bold=True, size=14, color=azul)
    resumen['A2'] = f'Periodo: {etiqueta_periodo(desde, hasta)}'
    if generado_por:
        resumen['A3'] = f'Generó: {generado_por}'
    resumen['A4'] = (
        'Solo movimientos con dinero real: ingresos pagados (de los parciales, '
        'lo cobrado), donativos no cancelados y egresos pagados.'
    )
    for celda in ('A2', 'A3', 'A4'):
        resumen[celda].font = Font(size=9, color='55607A')

    titulos = ['Mes', 'Servicios', 'Donativos', 'Total entradas', 'Salidas', 'Neto del mes', 'Acumulado']
    fila_titulos = 6
    for col, titulo in enumerate(titulos, start=1):
        celda = resumen.cell(row=fila_titulos, column=col, value=titulo)
        celda.font = encabezado
        celda.fill = relleno
        celda.alignment = Alignment(horizontal='center')

    filas = flujo_mensual(desde, hasta)
    fila = fila_titulos + 1
    for f in filas:
        valores = [f['etiqueta'], f['servicios'], f['donativos'], f['entradas'], f['salidas'], f['neto'], f['acumulado']]
        for col, valor in enumerate(valores, start=1):
            celda = resumen.cell(row=fila, column=col, value=valor)
            celda.border = borde_abajo
            if col > 1:
                celda.number_format = moneda
        fila += 1

    if filas:
        # El acumulado no se suma: es un saldo, así que el total lleva el
        # último valor de la columna, no la suma de los meses.
        claves = ['servicios', 'donativos', 'entradas', 'salidas', 'neto']
        resumen.cell(row=fila, column=1, value='TOTAL DEL PERIODO').font = Font(bold=True, color=azul)
        for col, clave in enumerate(claves, start=2):
            celda = resumen.cell(row=fila, column=col, value=sum((f[clave] for f in filas), Decimal('0')))
            celda.font = Font(bold=True, color=azul)
            celda.number_format = moneda
        celda = resumen.cell(row=fila, column=7, value=filas[-1]['acumulado'])
        celda.font = Font(bold=True, color=azul)
        celda.number_format = moneda
    else:
        resumen.cell(row=fila, column=1, value='No hay movimientos en el periodo seleccionado.')

    resumen.column_dimensions['A'].width = 22
    for col in range(2, 8):
        resumen.column_dimensions[get_column_letter(col)].width = 16

    detalle = libro.create_sheet('Movimientos')
    titulos = ['Fecha', 'Flujo', 'Tipo', 'Unidad', 'Concepto', 'Persona', 'Estatus', 'Monto']
    for col, titulo in enumerate(titulos, start=1):
        celda = detalle.cell(row=1, column=col, value=titulo)
        celda.font = encabezado
        celda.fill = relleno
    for indice, m in enumerate(movimientos_de_efectivo(desde, hasta), start=2):
        valores = [m['fecha'], m['flujo'], m['tipo'], m['unidad'], m['concepto'], m['persona'], m['estatus'], m['monto']]
        for col, valor in enumerate(valores, start=1):
            celda = detalle.cell(row=indice, column=col, value=valor)
            if col == 1:
                celda.number_format = 'dd/mm/yyyy'
            elif col == 8:
                celda.number_format = moneda
    detalle.freeze_panes = 'A2'
    for col, ancho in enumerate([12, 10, 11, 11, 40, 26, 13, 15], start=1):
        detalle.column_dimensions[get_column_letter(col)].width = ancho

    buffer = BytesIO()
    libro.save(buffer)
    respuesta = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    respuesta['Content-Disposition'] = (
        f'attachment; filename="flujo_efectivo_{sufijo_archivo(desde, hasta)}.xlsx"'
    )
    return respuesta


# ===== Concentrado de donativos =====

def concentrado_donativos(desde=None, hasta=None):
    """Donativos del periodo con sus totales. A diferencia de los otros dos
    reportes, aquí sí se listan los cancelados (marcados como tales): el
    concentrado sirve para revisar los CFDI emitidos, no solo el dinero. Los
    totales, en cambio, siguen excluyéndolos."""
    donativos = list(en_rango(Donativo.objects.all(), desde, hasta).order_by('fecha'))
    vigentes = [d for d in donativos if d.estatus_cfdi != Donativo.EstatusCFDI.CANCELADO]
    return {
        'donativos': donativos,
        'total': sum((d.monto for d in vigentes), Decimal('0')),
        'total_monetario': sum(
            (d.monto for d in vigentes if d.tipo == Donativo.Tipo.MONETARIO), Decimal('0')),
        'total_especie': sum(
            (d.monto for d in vigentes if d.tipo == Donativo.Tipo.ESPECIE), Decimal('0')),
        'total_cancelado': sum(
            (d.monto for d in donativos if d.estatus_cfdi == Donativo.EstatusCFDI.CANCELADO), Decimal('0')),
        'cuenta': len(donativos),
        'cuenta_vigentes': len(vigentes),
        'cuenta_sin_folio': len([d for d in vigentes if not d.folio_cfdi]),
        'donantes': len({d.donante_nombre.strip().lower() for d in vigentes}),
    }
